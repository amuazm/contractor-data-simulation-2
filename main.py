import shutil
from openpyxl import Workbook, load_workbook
import datetime
import numpy as np
from dateutil.relativedelta import relativedelta
import collections
# import matplotlib.pyplot as plt

import openpyxl

# Copy dataset to refer from as backup
src = "./Files/Input/Budget Dataset Modified.xlsx"
dest = "./Files/Output/Data.xlsx"
shutil.copyfile(src, dest)

# Create new empty workbook to start from
wb_result = Workbook()
# Sheets
wb_result.active.title = "Project Details"
ws_project_details = wb_result["Project Details"]
wb_result.create_sheet("Budget")
ws_budget = wb_result["Budget"]
wb_result.create_sheet("Categories")
ws_categories = wb_result["Categories"]
wb_result.create_sheet("Cash Outflow")
ws_cash_outflow = wb_result["Cash Outflow"]
wb_result.create_sheet("Reports")
ws_reports = wb_result["Reports"]

# Project Details
if True:
    # Load dataset to refer from
    wb_data = load_workbook(dest)
    # Copy entire Project Details sheet row by row from "data" to "result"
    for row in wb_data["Project Details"].iter_rows():
        l = []
        for cell in row:
            l.append(cell.value)
        ws_project_details.append(l)
    wb_result.close()
    # Styles
    for row in ws_project_details.iter_rows(min_row=2):
        row[11].style = "Currency"

# Project ID list
project_ids = ws_project_details["A"][1:]
project_ids = [i.value for i in project_ids]
# Overall budgets dict
overall_budgets = {}
for row in ws_project_details.iter_rows(min_row=2):
    overall_budgets[row[0].value] = row[11].value

# Budget
if True:
    # Headers
    ws_budget.append(["Project ID", "Start Date", "Duration (Months)", "Budgeted Cost per Month", "Overall Budget", "Projected Income"])

    i = 0
    for row in ws_budget.iter_rows(min_row=2, max_row=len(project_ids) + 1, max_col=6):
        # Project IDs
        row[0].value = project_ids[i]
        i += 1

        # Start Date
        # Generate a random date from 2 years ago to 6 months ago (for 6-24 months of data for each project)
        start = datetime.date.today() - relativedelta(years=2)
        end = datetime.date.today() - relativedelta(months=6)
        delta = end - start
        r = np.random.randint(0, delta.days + 1)
        new = start + datetime.timedelta(days=r)
        row[1].value = new

        # Duration (Months)
        # 2.5 to 5 years
        row[2].value = np.random.randint(30, 60)

        # Cost per Month
        row[3].value = overall_budgets[row[0].value] / row[2].value
        row[3].style = "Currency"

        # Overall Budget
        row[4].value = overall_budgets[row[0].value]
        row[4].style = "Currency"

        # Contract Pay
        row[5].value = overall_budgets[row[0].value] * np.random.randint(1020, 1080)/1000
        row[5].style = "Currency"


# Categories
if True:
    # Headers
    ws_categories.append(["Project ID", "Category", "Category Budget"])

    categories = [
        'Direct Labour',
        'Supplied Labour',
        'Sub-contractor',
        'Other Materials',
        'Small Tools & Safety Item',
        'Other Consumable',
        'Transportation',
        'Repair & Maintenance',
        'Site Office Expense',
        'Food, Refreshment & Entertainment',
        'Travelling & Vehicles',
        'Main Steel Materials',
        'Stainless Steel Materials',
        'Aluminium Materials',
        'Equipment',
        'Supervision',
        'Insurance'
    ]

    # Randomly spread out percentages of how much each category is budgeted. Maximum of 100%. Percentages can be found in k_nums.
    for id in project_ids:
        n, k = overall_budgets[id], len(categories)
        vals = np.random.default_rng().dirichlet(np.ones(k), size=1)
        k_nums = [round(v) for v in vals[0]*n]
        i = 0
        for category in categories:
            ws_categories.append([id, category, k_nums[i]])
            i += 1

# Start dates dict
start_dates = {}
for row in ws_budget.iter_rows(min_row=2):
    start_dates[row[0].value] = row[1].value
# Project durations dict
project_durations = {}
for row in ws_budget.iter_rows(min_row=2):
    project_durations[row[0].value] = row[2].value
# Budgets by category nested dict
budgets_by_cat = {}
for row in ws_categories.iter_rows(min_row=2):
    if row[0].value not in budgets_by_cat:
        budgets_by_cat[row[0].value] = {}
    budgets_by_cat[row[0].value][row[1].value] = row[2].value

# Cash Outflow
if True:
    # Headers
    ws_cash_outflow.append(["Project ID", "Date", "Category", "Actual Category Monthly Cost"])

    for id in project_ids:
        cash_outflow_date = start_dates[id]
        # First cash outflow date will be 1 month after project start date
        cash_outflow_date += relativedelta(months=1)
        # Add 1 month to date until today's date
        # TODO: 80.0% to 120.0% budget multiplier
        while(datetime.date.today() - cash_outflow_date > datetime.timedelta(days=0)):
            for category in categories:
                # TODO: 95% to 105% category budget multiplier
                # Actual Category Monthly Cost = Category Budget / Project Duration
                monthly_cash_outflow = budgets_by_cat[id][category] / project_durations[id]
                ws_cash_outflow.append([id, cash_outflow_date, category, monthly_cash_outflow])
            cash_outflow_date += relativedelta(months=1)

# Actual costs by date and category nested dict
# This is a bit overkill. Maybe not.
actual_by_cat_date = {}
for row in ws_cash_outflow.iter_rows(min_row=2):
    if row[0].value not in actual_by_cat_date:
        actual_by_cat_date[row[0].value] = {}
    if row[1].value not in actual_by_cat_date[row[0].value]:
        actual_by_cat_date[row[0].value][row[1].value] = {}
    actual_by_cat_date[row[0].value][row[1].value][row[2].value] = row[3].value
# Monthly budgets dict
monthly_budgets = {}
for row in ws_budget.iter_rows(min_row=2):
    monthly_budgets[row[0].value] = row[3].value

# Reports
if True:
    # Headers
    ws_reports.append(["Project ID", "Date", "Completion", "ACWP", "BCWP", "BCWS"])

    # Project ID
    for id in project_ids:
        months_passed = 1
        acwp = 0
        # Date of Report
        for date in actual_by_cat_date[id]:
            # Completion
            completion = months_passed * (1/project_durations[id])

            # ACWP
            date_cost = 0
            for category in actual_by_cat_date[id][date]:
                    date_cost += actual_by_cat_date[id][date][category]
            acwp += date_cost

            # BCWP
            bcwp = overall_budgets[id] * completion

            # BCWS
            bcws = months_passed * monthly_budgets[id]

            ws_reports.append([id, date, completion, acwp, bcwp, bcws])
            months_passed += 1

    # Styles
    for row in ws_reports.iter_rows(min_row=2):
        row[2].style = "Percent"
        row[3].style = "Currency"
        row[4].style = "Currency"
        row[5].style = "Currency"

# Save result
wb_result.save("./Files/Output/Result.xlsx")